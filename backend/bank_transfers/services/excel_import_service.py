"""
Excel import service layer
Handles business logic for Excel file processing and import
"""
import openpyxl
from ..hungarian_account_validator import validate_and_format_hungarian_account_number
from .beneficiary_service import BeneficiaryService


class ExcelImportService:
    """Service for Excel import business logic"""
    
    @staticmethod
    def import_beneficiaries_from_excel(excel_file, company):
        """Import beneficiaries from Excel file with proper validation and error handling"""
        try:
            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active
        except Exception as e:
            raise ValueError(f"Cannot read Excel file: {str(e)}")
        
        beneficiaries = []
        errors = []
        
        # Detect header row and start from the row after headers
        start_row = 3
        header_keywords = ['név', 'name', 'számlaszám', 'account', 'összeg', 'amount']
        
        # Check if row 3 contains headers and skip if so
        if worksheet.max_row >= 3:
            row_3_values = [str(cell.value or '').lower().strip() for cell in worksheet[3][:6]]
            if any(keyword in ' '.join(row_3_values) for keyword in header_keywords):
                start_row = 4  # Skip header row and start from row 4
        
        for row_num, row in enumerate(worksheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
            if not any(row[:6]):
                continue
            
            try:
                comment, name, account_number, amount, exec_date, remittance = row[:6]
                
                # Convert to strings and strip whitespace
                name = str(name or '').strip() if name is not None else ''
                account_number = str(account_number or '').strip() if account_number is not None else ''
                
                # Skip if either name or account number is missing
                if not name or not account_number:
                    if name or account_number:  # Only log if partially filled
                        errors.append(f'Row {row_num}: Missing name or account number')
                    continue
                
                # Skip obvious header rows
                if name.lower() in ['név', 'name'] or account_number.lower() in ['számlaszám', 'account']:
                    continue
                
                # Validate and format Hungarian account number
                validation_result = validate_and_format_hungarian_account_number(
                    account_number, validate_checksum=False
                )
                
                if not validation_result.is_valid:
                    error_msg = f'Row {row_num}: Érvénytelen számlaszám "{account_number}": {validation_result.error}'
                    errors.append(error_msg)
                    continue
                
                # Use the properly formatted account number
                formatted_account_number = validation_result.formatted
                
                # Create beneficiary using service
                beneficiary, created = BeneficiaryService.find_or_create_from_excel_data(
                    company=company,
                    name=name,
                    account_number=formatted_account_number,
                    description=str(comment or '').strip(),
                    remittance_information=str(remittance or '').strip()
                )
                
                if created:
                    beneficiaries.append(beneficiary)
                
            except Exception as e:
                error_msg = f"Row {row_num}: {str(e)}"
                errors.append(error_msg)
                continue
        
        return {
            'beneficiaries': beneficiaries,
            'errors': errors,
            'imported_count': len(beneficiaries)
        }
    
    @staticmethod
    def validate_excel_file(excel_file):
        """Validate Excel file format and structure"""
        try:
            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active
            
            if worksheet.max_row < 3:
                raise ValueError("Excel file must have at least 3 rows")
            
            return True
        except Exception as e:
            raise ValueError(f"Invalid Excel file: {str(e)}")