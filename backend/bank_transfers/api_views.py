from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FileUploadParser
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db import transaction, models
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
import json
from datetime import date

from .models import BankAccount, Beneficiary, TransferTemplate, TemplateBeneficiary, Transfer, TransferBatch
from .serializers import (
    BankAccountSerializer, BeneficiarySerializer, TransferTemplateSerializer,
    TemplateBeneficiarySerializer, TransferSerializer, TransferBatchSerializer,
    TransferCreateFromTemplateSerializer, BulkTransferSerializer,
    ExcelImportSerializer, XMLGenerateSerializer
)
from .utils import generate_xml
from .pdf_processor import PDFTransactionProcessor
from .kh_export import KHBankExporter

# Health check endpoint for Railway
def health_check(request):
    """Simple health check endpoint for Railway deployment"""
    return JsonResponse({
        'status': 'healthy',
        'timestamp': timezone.now().isoformat(),
        'service': 'transferXMLGenerator-backend'
    })

class BankAccountViewSet(viewsets.ModelViewSet):
    """
    Bank számlák kezelése
    
    list: Az összes bank számla listája
    create: Új bank számla létrehozása
    retrieve: Egy konkrét bank számla adatai
    update: Bank számla módosítása
    destroy: Bank számla törlése
    """
    queryset = BankAccount.objects.all()
    serializer_class = BankAccountSerializer
    
    @swagger_auto_schema(
        operation_description="Alapértelmezett bank számla lekérése",
        responses={200: BankAccountSerializer, 404: 'Nincs alapértelmezett számla'}
    )
    @action(detail=False, methods=['get'])
    def default(self, request):
        """Alapértelmezett bank számla lekérése"""
        account = BankAccount.objects.filter(is_default=True).first()
        if account:
            serializer = self.get_serializer(account)
            return Response(serializer.data)
        return Response({'detail': 'No default account found'}, status=404)

class BeneficiaryViewSet(viewsets.ModelViewSet):
    """
    Kedvezményezettek kezelése
    
    Támogatott szűrések:
    - is_active: true/false
    - is_frequent: true/false  
    - search: név alapján keresés
    """
    queryset = Beneficiary.objects.all()
    serializer_class = BeneficiarySerializer
    
    def get_queryset(self):
        queryset = Beneficiary.objects.all()
        
        # Filtering
        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
            
        is_frequent = self.request.query_params.get('is_frequent', None)
        if is_frequent is not None:
            queryset = queryset.filter(is_frequent=is_frequent.lower() == 'true')
            
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(name__icontains=search)
            
        return queryset
    
    @swagger_auto_schema(
        operation_description="Gyakori kedvezményezettek listája",
        responses={200: BeneficiarySerializer(many=True)}
    )
    @action(detail=False, methods=['get'])
    def frequent(self, request):
        """Gyakori kedvezményezettek listája"""
        beneficiaries = Beneficiary.objects.filter(is_frequent=True, is_active=True)
        serializer = self.get_serializer(beneficiaries, many=True)
        return Response(serializer.data)
    
    @swagger_auto_schema(
        operation_description="Kedvezményezett gyakori státuszának váltása",
        responses={200: BeneficiarySerializer}
    )
    @action(detail=True, methods=['post'])
    def toggle_frequent(self, request, pk=None):
        """Gyakori státusz váltása"""
        beneficiary = self.get_object()
        beneficiary.is_frequent = not beneficiary.is_frequent
        beneficiary.save()
        serializer = self.get_serializer(beneficiary)
        return Response(serializer.data)

class TransferTemplateViewSet(viewsets.ModelViewSet):
    """
    Utalási sablonok kezelése
    
    A sablonok lehetővé teszik gyakori utalási ciklusok (pl. hó eleji fizetések)
    gyors betöltését és ismételt használatát.
    """
    queryset = TransferTemplate.objects.filter(is_active=True).order_by('-created_at')
    serializer_class = TransferTemplateSerializer
    
    @swagger_auto_schema(
        operation_description="Kedvezményezett hozzáadása sablonhoz",
        request_body=TemplateBeneficiarySerializer,
        responses={201: TemplateBeneficiarySerializer, 400: 'Validation error'}
    )
    @action(detail=True, methods=['post'])
    def add_beneficiary(self, request, pk=None):
        """Kedvezményezett hozzáadása sablonhoz"""
        template = self.get_object()
        serializer = TemplateBeneficiarySerializer(data=request.data)
        
        if serializer.is_valid():
            serializer.save(template=template)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @swagger_auto_schema(
        operation_description="Kedvezményezett eltávolítása sablonból",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                'beneficiary_id': openapi.Schema(type=openapi.TYPE_INTEGER, description='Kedvezményezett ID')
            }
        ),
        responses={204: 'Successfully removed', 404: 'Beneficiary not found'}
    )
    @action(detail=True, methods=['delete'])
    def remove_beneficiary(self, request, pk=None):
        """Kedvezményezett eltávolítása sablonból"""
        template = self.get_object()
        beneficiary_id = request.data.get('beneficiary_id')
        
        try:
            template_beneficiary = TemplateBeneficiary.objects.get(
                template=template, 
                beneficiary_id=beneficiary_id
            )
            template_beneficiary.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except TemplateBeneficiary.DoesNotExist:
            return Response({'detail': 'Beneficiary not found in template'}, status=404)
    
    @swagger_auto_schema(
        operation_description="Sablon kedvezményezett frissítése",
        request_body=TemplateBeneficiarySerializer,
        responses={
            200: TemplateBeneficiarySerializer,
            404: 'Template beneficiary not found'
        }
    )
    @action(detail=True, methods=['put'])
    def update_beneficiary(self, request, pk=None):
        """Sablon kedvezményezett frissítése (összeg, közlemény, sorrend)"""
        template = self.get_object()
        beneficiary_id = request.data.get('beneficiary_id')
        
        try:
            template_beneficiary = TemplateBeneficiary.objects.get(
                template=template, 
                beneficiary_id=beneficiary_id
            )
            serializer = TemplateBeneficiarySerializer(template_beneficiary, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except TemplateBeneficiary.DoesNotExist:
            return Response({'detail': 'Beneficiary not found in template'}, status=404)
    
    @swagger_auto_schema(
        operation_description="Sablon betöltése utalások létrehozásához",
        request_body=TransferCreateFromTemplateSerializer,
        responses={
            200: openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    'template': openapi.Schema(type=openapi.TYPE_OBJECT),
                    'transfers': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_OBJECT))
                }
            )
        }
    )
    @action(detail=True, methods=['post'])
    def load_transfers(self, request, pk=None):
        """Sablon betöltése és utalás adatok generálása"""
        template = self.get_object()
        serializer = TransferCreateFromTemplateSerializer(data=request.data)
        
        if serializer.is_valid():
            data = serializer.validated_data
            originator_account = get_object_or_404(BankAccount, id=data['originator_account_id'])
            execution_date = data['execution_date']
            
            transfers = []
            for template_beneficiary in template.template_beneficiaries.filter(is_active=True):
                # Use template's default execution date if available, otherwise use user-provided date
                beneficiary_execution_date = (
                    template_beneficiary.default_execution_date.strftime('%Y-%m-%d') 
                    if template_beneficiary.default_execution_date 
                    else execution_date
                )
                
                transfer_data = {
                    'originator_account': originator_account.id,
                    'beneficiary': template_beneficiary.beneficiary.id,
                    'amount': template_beneficiary.default_amount or 0,
                    'currency': 'HUF',
                    'execution_date': beneficiary_execution_date,
                    'remittance_info': template_beneficiary.default_remittance or template_beneficiary.beneficiary.remittance_information or '',
                    'template': template.id
                }
                transfers.append(transfer_data)
            
            return Response({
                'template': TransferTemplateSerializer(template).data,
                'transfers': transfers
            })
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @swagger_auto_schema(
        operation_description="PDF fájlokból sablon létrehozása/frissítése",
        manual_parameters=[
            openapi.Parameter('pdf_files', openapi.IN_FORM, type=openapi.TYPE_FILE, required=True, description='PDF fájlok (több is)'),
            openapi.Parameter('template_name', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Sablon neve (opcionális)'),
            openapi.Parameter('template_id', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False, description='Meglévő sablon ID frissítéshez (opcionális)')
        ],
        responses={
            200: openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    'template': openapi.Schema(type=openapi.TYPE_OBJECT),
                    'transactions_processed': openapi.Schema(type=openapi.TYPE_INTEGER),
                    'beneficiaries_matched': openapi.Schema(type=openapi.TYPE_INTEGER),
                    'beneficiaries_created': openapi.Schema(type=openapi.TYPE_INTEGER),
                    'consolidations': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_STRING)),
                    'preview': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_OBJECT)),
                    'total_amount': openapi.Schema(type=openapi.TYPE_NUMBER)
                }
            ),
            400: 'Hibás PDF vagy feldolgozási hiba'
        }
    )
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser])
    def process_pdf(self, request):
        """
        PDF fájlokból tranzakciók kinyerése és sablon létrehozása
        
        Támogatott formátumok:
        - NAV adó és járulék befizetési PDF-ek
        - Banki utalások / fizetési lista PDF-ek
        
        A rendszer automatikusan:
        - Felismeri a meglévő kedvezményezetteket (nem hoz létre duplikátumokat)
        - Összevonja az azonos számlára érkező átutalásokat
        - Létrehoz új sablont vagy frissíti a meglévőt
        """
        try:
            pdf_files = request.FILES.getlist('pdf_files')
            template_name = request.data.get('template_name')
            template_id = request.data.get('template_id')
            
            if not pdf_files:
                return Response({
                    'error': 'Nem található PDF fájl',
                    'details': 'Legalább egy PDF fájlt fel kell tölteni'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate PDF files
            for pdf_file in pdf_files:
                if not pdf_file.name.lower().endswith('.pdf'):
                    return Response({
                        'error': f'Hibás fájl formátum: {pdf_file.name}',
                        'details': 'Csak PDF fájlok engedélyezettek'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Process PDFs
            processor = PDFTransactionProcessor()
            result = processor.process_pdf_files(
                pdf_files=pdf_files,
                template_name=template_name,
                template_id=int(template_id) if template_id else None
            )
            
            return Response(result, status=status.HTTP_200_OK)
            
        except ValueError as e:
            return Response({
                'error': 'PDF feldolgozási hiba',
                'details': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            return Response({
                'error': 'Váratlan hiba történt',
                'details': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class TransferViewSet(viewsets.ModelViewSet):
    """
    Utalások kezelése
    
    Támogatott szűrések:
    - is_processed: true/false
    - template: sablon ID
    - execution_date_from: dátum (YYYY-MM-DD)
    - execution_date_to: dátum (YYYY-MM-DD)
    """
    queryset = Transfer.objects.all()
    serializer_class = TransferSerializer
    
    def get_queryset(self):
        queryset = Transfer.objects.select_related('beneficiary', 'originator_account').all()
        
        # Filtering
        is_processed = self.request.query_params.get('is_processed', None)
        if is_processed is not None:
            queryset = queryset.filter(is_processed=is_processed.lower() == 'true')
            
        template_id = self.request.query_params.get('template', None)
        if template_id:
            queryset = queryset.filter(template_id=template_id)
            
        execution_date_from = self.request.query_params.get('execution_date_from', None)
        if execution_date_from:
            queryset = queryset.filter(execution_date__gte=execution_date_from)
            
        execution_date_to = self.request.query_params.get('execution_date_to', None)
        if execution_date_to:
            queryset = queryset.filter(execution_date__lte=execution_date_to)
            
        return queryset
    
    @swagger_auto_schema(
        operation_description="Több utalás egyszerre létrehozása",
        request_body=BulkTransferSerializer,
        responses={201: 'Successfully created', 400: 'Validation error'}
    )
    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Tömeges utalás létrehozás"""
        serializer = BulkTransferSerializer(data=request.data)
        
        if serializer.is_valid():
            transfers_data = serializer.validated_data['transfers']
            batch_name = serializer.validated_data.get('batch_name')
            
            with transaction.atomic():
                transfers = []
                for transfer_data in transfers_data:
                    transfer_serializer = TransferSerializer(data=transfer_data)
                    if transfer_serializer.is_valid():
                        transfer = transfer_serializer.save()
                        transfers.append(transfer)
                    else:
                        return Response(transfer_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                
                # Create batch if name provided
                if batch_name and transfers:
                    batch = TransferBatch.objects.create(name=batch_name)
                    batch.transfers.set(transfers)
                    total_amount = sum(t.amount for t in transfers)
                    batch.total_amount = total_amount
                    batch.save()
            
            return Response({
                'transfers': TransferSerializer(transfers, many=True).data,
                'count': len(transfers)
            }, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @swagger_auto_schema(
        operation_description="XML generálás kiválasztott utalásokból",
        request_body=XMLGenerateSerializer,
        responses={
            200: openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    'xml': openapi.Schema(type=openapi.TYPE_STRING, description='Generált XML tartalom'),
                    'transfer_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                    'total_amount': openapi.Schema(type=openapi.TYPE_STRING)
                }
            ),
            404: 'No transfers found'
        }
    )
    @action(detail=False, methods=['post'])
    def generate_xml(self, request):
        """XML generálás kiválasztott utalásokból"""
        serializer = XMLGenerateSerializer(data=request.data)
        
        if serializer.is_valid():
            transfer_ids = serializer.validated_data['transfer_ids']
            batch_name = serializer.validated_data.get('batch_name')
            
            transfers = Transfer.objects.filter(id__in=transfer_ids).select_related(
                'beneficiary', 'originator_account'
            ).order_by('order', 'execution_date')
            
            if not transfers:
                return Response({'detail': 'No transfers found'}, status=404)
            
            # Generate XML
            xml_content = generate_xml(transfers)
            
            # Create batch if name provided
            if batch_name:
                # Get next order number
                max_order = TransferBatch.objects.aggregate(max_order=models.Max('order'))['max_order'] or 0
                
                batch = TransferBatch.objects.create(
                    name=batch_name,
                    xml_generated_at=timezone.now(),
                    total_amount=sum(t.amount for t in transfers),
                    order=max_order + 1  # Set incrementing order
                )
                batch.transfers.set(transfers)
            
            # Mark transfers as processed
            transfers.update(is_processed=True)
            
            return Response({
                'xml': xml_content,
                'transfer_count': len(transfers),
                'total_amount': str(sum(t.amount for t in transfers))
            })
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @swagger_auto_schema(
        operation_description="KH Bank formátumú szöveges export generálás",
        request_body=XMLGenerateSerializer,
        responses={
            200: openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    'content': openapi.Schema(type=openapi.TYPE_STRING, description='KH Bank .HUF.CSV file tartalma'),
                    'filename': openapi.Schema(type=openapi.TYPE_STRING, description='Ajánlott fájlnév'),
                    'transfer_count': openapi.Schema(type=openapi.TYPE_INTEGER),
                    'total_amount': openapi.Schema(type=openapi.TYPE_STRING)
                }
            ),
            404: 'No transfers found',
            400: 'Export error'
        }
    )
    @action(detail=False, methods=['post'])
    def generate_kh_export(self, request):
        """KH Bank formátumú .HUF.CSV export generálás"""
        serializer = XMLGenerateSerializer(data=request.data)
        
        if serializer.is_valid():
            transfer_ids = serializer.validated_data['transfer_ids']
            batch_name = serializer.validated_data.get('batch_name')
            
            transfers = Transfer.objects.filter(id__in=transfer_ids).select_related(
                'beneficiary', 'originator_account'
            ).order_by('order', 'execution_date')
            
            if not transfers:
                return Response({'detail': 'No transfers found'}, status=404)
            
            try:
                # Generate KH Bank export
                exporter = KHBankExporter()
                kh_content = exporter.generate_kh_export(transfers)
                filename = exporter.get_filename(batch_name)
                
                # Create batch if name provided
                if batch_name:
                    # Get next order number
                    max_order = TransferBatch.objects.aggregate(max_order=models.Max('order'))['max_order'] or 0
                    
                    batch = TransferBatch.objects.create(
                        name=f"{batch_name} (KH Export)",
                        xml_generated_at=timezone.now(),
                        total_amount=sum(t.amount for t in transfers),
                        order=max_order + 1
                    )
                    batch.transfers.set(transfers)
                
                # Mark transfers as processed
                transfers.update(is_processed=True)
                
                return Response({
                    'content': kh_content,
                    'filename': filename,
                    'transfer_count': len(transfers),
                    'total_amount': str(sum(t.amount for t in transfers))
                })
                
            except ValueError as e:
                return Response({
                    'detail': 'KH Bank export error',
                    'error': str(e)
                }, status=status.HTTP_400_BAD_REQUEST)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class TransferBatchViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Utalási kötegek megtekintése
    
    A kötegek automatikusan létrejönnek XML generáláskor,
    és nyilvántartják a feldolgozott utalásokat.
    """
    queryset = TransferBatch.objects.all()
    serializer_class = TransferBatchSerializer
    
    @swagger_auto_schema(
        operation_description="XML fájl letöltése köteghez",
        responses={200: 'XML file', 404: 'No transfers in batch'}
    )
    @action(detail=True, methods=['get'])
    def download_xml(self, request, pk=None):
        """XML fájl letöltése - regenerálja az XML-t a mentett adatokból"""
        batch = self.get_object()
        transfers = batch.transfers.select_related('beneficiary', 'originator_account').order_by('order', 'execution_date')
        
        if not transfers:
            return Response({'detail': 'No transfers in batch'}, status=404)
        
        # Regenerate XML from saved transfer data
        xml_content = generate_xml(transfers)
        
        response = HttpResponse(xml_content, content_type='application/xml')
        response['Content-Disposition'] = f'attachment; filename="{batch.xml_filename}"'
        return response
    
    @swagger_auto_schema(
        operation_description="XML köteg megjelölése bankban felhasználtként",
        responses={200: 'Batch marked as used', 404: 'Batch not found'}
    )
    @action(detail=True, methods=['post'])
    def mark_used_in_bank(self, request, pk=None):
        """XML köteg megjelölése bankban felhasználtként"""
        batch = self.get_object()
        batch.used_in_bank = True
        batch.bank_usage_date = timezone.now()
        batch.save()
        
        return Response({
            'detail': 'Batch marked as used in bank',
            'used_in_bank': batch.used_in_bank,
            'bank_usage_date': batch.bank_usage_date
        })
    
    @swagger_auto_schema(
        operation_description="XML köteg megjelölése nem felhasználtként",
        responses={200: 'Batch marked as unused', 404: 'Batch not found'}
    )
    @action(detail=True, methods=['post'])
    def mark_unused_in_bank(self, request, pk=None):
        """XML köteg megjelölése nem felhasználtként"""
        batch = self.get_object()
        batch.used_in_bank = False
        batch.bank_usage_date = None
        batch.save()
        
        return Response({
            'detail': 'Batch marked as unused in bank',
            'used_in_bank': batch.used_in_bank,
            'bank_usage_date': batch.bank_usage_date
        })

class ExcelImportView(APIView):
    """
    Excel fájl import kedvezményezettek tömeges feltöltéséhez
    
    Támogatott formátum:
    - 3. sortól kezdődnek az adatok
    - Oszlopok: Megjegyzés, Név, Számlaszám, Összeg, Dátum, Közlemény
    - Csak a Név és Számlaszám kötelező
    """
    parser_classes = [MultiPartParser]
    
    @swagger_auto_schema(
        operation_description="Excel fájl feltöltése kedvezményezettek importálásához",
        request_body=ExcelImportSerializer,
        responses={
            200: openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    'message': openapi.Schema(type=openapi.TYPE_STRING),
                    'beneficiaries': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_OBJECT))
                }
            ),
            400: 'Import error'
        }
    )
    def post(self, request):
        serializer = ExcelImportSerializer(data=request.data)
        
        if serializer.is_valid():
            excel_file = serializer.validated_data['file']
            import_type = serializer.validated_data['import_type']
            
            try:
                if import_type == 'beneficiaries':
                    beneficiaries = self.import_beneficiaries_from_excel(excel_file)
                    return Response({
                        'message': f'{len(beneficiaries)} kedvezményezett importálva',
                        'beneficiaries': BeneficiarySerializer(beneficiaries, many=True).data
                    })
                else:
                    return Response({'detail': 'Invalid import type'}, status=400)
                    
            except Exception as e:
                return Response({'detail': str(e)}, status=400)
        
        return Response(serializer.errors, status=400)
    
    def import_beneficiaries_from_excel(self, excel_file):
        """Kedvezményezettek importálása Excel fájlból"""
        import openpyxl
        
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active
        
        beneficiaries = []
        
        # Detect header row and start from the row after headers
        start_row = 3
        header_keywords = ['név', 'name', 'számlaszám', 'account', 'összeg', 'amount']
        
        # Check if row 3 contains headers and skip if so
        if worksheet.max_row >= 3:
            row_3_values = [str(cell.value or '').lower().strip() for cell in worksheet[3][:6]]
            if any(keyword in ' '.join(row_3_values) for keyword in header_keywords):
                start_row = 4  # Skip header row and start from row 4
        
        for row in worksheet.iter_rows(min_row=start_row, values_only=True):
            if not any(row[:6]):
                continue
                
            try:
                comment, name, account_number, amount, exec_date, remittance = row[:6]
                
                if not all([name, account_number]):
                    continue
                
                beneficiary, created = Beneficiary.objects.get_or_create(
                    name=str(name).strip(),
                    account_number=str(account_number).strip(),
                    defaults={
                        'description': '',
                        'is_active': True,
                        'remittance_information': str(comment or '').strip()
                    }
                )
                
                if created:
                    beneficiaries.append(beneficiary)
                    
            except Exception as e:
                print(f"Error processing row: {e}")
                continue
        
        return beneficiaries